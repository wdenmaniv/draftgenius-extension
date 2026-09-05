#!/usr/bin/env python3
"""Converts one of elboberto's yearly draft-prep spreadsheets into the JSON
shape scripts/analyze-historical-bias.js expects — completing the "upload a
new year, rerun the analysis" process for the OTHER half (analyze-historical-
bias.js only auto-discovers years that already have a
data/historical-projections/<year>/ directory; this script is what creates
one from the raw .xlsm elboberto actually sends).

Usage:
    python3 scripts/import-elboberto-projections.py <year> <path-to-xlsm>

Reads the six `*_Raw` sheets (QB_Raw, RB_Raw, WR_Raw, TE_Raw, K_Raw, DEF_Raw)
— confirmed, not guessed, to carry the exact same column shape across every
year checked so far (2023 verified against this script; matches the
already-committed 2022/2024/2025 JSON files field-for-field) — and writes
data/historical-projections/<year>/{qb,rb,wr,te,k,def}.json, one array of
per-player dicts per position, keyed exactly as elboberto's own headers
(e.g. "PASSING YDS", "MISC FL") so historicalErrorAdjustments.js's consumer
(analyze-historical-bias.js) needs no per-year special-casing.

Requires openpyxl (already used elsewhere in this repo's tooling).
"""
import json
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent

SHEETS = {
    "qb": "QB_Raw",
    "rb": "RB_Raw",
    "wr": "WR_Raw",
    "te": "TE_Raw",
    "k": "K_Raw",
    "def": "DEF_Raw",
}


def sheet_to_rows(ws):
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter)
    out = []
    for row in rows_iter:
        record = dict(zip(header, row))
        player = record.get("Player")
        if not player or not str(player).strip():
            continue
        out.append(record)
    return out


def main():
    if len(sys.argv) != 3:
        print(f"Usage: {sys.argv[0]} <year> <path-to-xlsm>", file=sys.stderr)
        sys.exit(1)
    year = sys.argv[1]
    xlsm_path = Path(sys.argv[2]).expanduser()
    if not xlsm_path.exists():
        print(f"File not found: {xlsm_path}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsm_path, data_only=True, read_only=True)
    missing = [name for name in SHEETS.values() if name not in wb.sheetnames]
    if missing:
        print(f"Missing expected sheet(s) in {xlsm_path.name}: {missing}", file=sys.stderr)
        print(f"Sheets found: {wb.sheetnames}", file=sys.stderr)
        sys.exit(1)

    out_dir = ROOT / "data/historical-projections" / year
    out_dir.mkdir(parents=True, exist_ok=True)

    for fname, sheet_name in SHEETS.items():
        rows = sheet_to_rows(wb[sheet_name])
        out_path = out_dir / f"{fname}.json"
        out_path.write_text(json.dumps(rows, indent=2))
        print(f"  {fname}: {len(rows)} players -> {out_path.relative_to(ROOT)}")

    print(f"\nWrote {out_dir.relative_to(ROOT)}/*.json from {xlsm_path.name}")
    print("Next: node scripts/analyze-historical-bias.js  (auto-discovers this year now that both projections and actuals exist)")


if __name__ == "__main__":
    main()
